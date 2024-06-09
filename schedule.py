import os
import openpyxl
from openpyxl.styles import Font
from openpyxl .styles.colors import Color
import numpy as np

class Schedule_EDIT():
    def __init__(self):
        self.input_file_name = 'original.xlsx'
        self.output_file_name = 'result.xlsx'
        self.wb = None
        self.sheet = None
        self.data = None
        self.data_leader = None
        self.set_begin_leader = [4,3]
        self.set_end_leader = [7,33]
        self.set_begin = [8,3]
        self.set_end = [19,33]
        self.set_worker_row = []
        self.set_even_odd = -1
        self.ref_sort = ['N','A','B','A','B','1','2','3','4','5','6','7','8','9','10']
        self.new_data = []
        
    def read_xlsx(self):
        self.wb = openpyxl.load_workbook(self.input_file_name)
        self.sheet = self.wb['工作表1']

    def get_values(self):
        arr = []
        arr_leader = []
        get_worker_num = False
        for r,row in enumerate(self.sheet):
            arr2 = []
            arr2_leader = []
            for c,col in enumerate(row):
                if c >= self.set_begin[1] and c <= self.set_end[1]:
                    arr2.append(col.value)
                if c >= self.set_begin_leader[1] and c <= self.set_end_leader[1]:
                    arr2_leader.append(col.value)

                if col.value == '應上班人數':
                    get_worker_num = True
                elif get_worker_num:
                    self.set_worker_row.append(col.value)
                
                    
                
            if r >= self.set_begin[0] and r <= self.set_end[0]:
                arr.append(arr2)
            if r >= self.set_begin_leader[0] and r <= self.set_end_leader[0]:
                arr_leader.append(arr2_leader)


        self.data = np.array(arr)
        self.data_leader = np.array(arr_leader)
        print(self.set_worker_row)
        print(self.data_leader)

    def save_data(self):
        wb_out = openpyxl.load_workbook(self.input_file_name)
        sheet = wb_out['工作表1']

        for r,row in enumerate(self.new_data):
            for c,cell in enumerate(row):
                sheet.cell(r+self.set_begin[0]+1,c+self.set_begin[1]+1).value = cell
                if cell == 'H' or cell == 'X':
                    sheet.cell(r+self.set_begin[0]+1,c+self.set_begin[1]+1).font = Font(u'標楷體',color = "FFFF0000")
        wb_out.save(self.output_file_name)

    def create_ref(self):
        max_r,max_c = [self.set_end[i] - self.set_begin[i] for i in range(2)]
        max_r += 1
        max_c += 1
        self.new_data = np.zeros([max_r,max_c],dtype=str)
        set_N = -1

        for c in range(max_c):
            if c % 2 == self.set_even_odd:
                r = 0
                start_fill = False
                fill_count = 0

                if set_N != -1:
                    find_N_r = set_N
                    while True:
                        if self.data[find_N_r,c] != 'X':
                            self.new_data[find_N_r,c] = 'N'
                            break
                        else:
                            if find_N_r != 0:
                                find_N_r -= 1
                            else:
                                find_N_r = max_r - 1

                while True:
                    if self.data[r,c] == 'N' and not start_fill:
                        set_N = r
                        fill_count = 0
                        start_fill = True
                        
                    if self.new_data[r,c] == 'N':
                        if not start_fill:
                            fill_count = 0
                            start_fill = True
                        else:
                            if set_N != 0:
                                set_N -= 1
                            else :
                                set_N = max_r -1 
                            break

                    if start_fill:
                        if self.data[r,c] == 'X':
                            self.new_data[r,c] = 'X'
                        else:
                            self.new_data[r,c] = self.ref_sort[fill_count]
                            fill_count += 1
                    r += 1
                    if r == max_r and start_fill:
                        if start_fill:
                            r = 0
                        else:
                            print('Cannot find N')

    def find_max(self,get_list):
        get_list = np.array(list(get_list))
        max_index = np.argmax(get_list)
        max_val = get_list[max_index]
        if max_val != 'N' and max_val != 'A' and max_val != 'B' and max_val != 'H' and max_val != 'X':
            return max_index
        else:
            get_list[max_index] = 0
            return self.find_max(get_list)
        
    def get_sort_element(self,get_list):
        if 'B' in get_list:
            get_list.remove('B')
            return 'B'
        elif 'A' in get_list:
            get_list.remove('A')
            return 'A'
        elif 'N' in get_list:
            get_list.remove('N')
            return 'N'
        else:
            max_val = max(get_list)
            get_list.remove(max_val)
            return max_val

        
                    
    def collate_data(self):
        max_r,max_c = [self.set_end[i] - self.set_begin[i] for i in range(2)]
        max_r += 1
        max_c += 1

        for c in range(max_c):
            if c % 2 == self.set_even_odd:
                absent_work = []
                for r in range(max_r):
                    if self.data[r,c] == 'H':
                        absent_work.append(self.new_data[r,c])
                        self.new_data[r,c] = 'H'

                print(self.new_data[self.find_max(self.new_data[:,c]),c])
                print(absent_work)

                while absent_work != []:
                    val = self.get_sort_element(get_list = absent_work)
                    if val == 'N' or val == 'A' or val == 'B':
                        self.new_data[self.find_max(self.new_data[:,c]),c] = val
                    elif val < self.new_data[self.find_max(self.new_data[:,c]),c]:
                        self.new_data[self.find_max(self.new_data[:,c]),c] = val
                print(self.set_worker_row[c])

if __name__ == '__main__':
    sch = Schedule_EDIT()
    sch.set_even_odd = 1    # odd = 0 even = 1
    sch.read_xlsx()
    sch.get_values()
    sch.create_ref()
    sch.collate_data()
    sch.save_data()

